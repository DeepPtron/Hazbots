from flask import Flask, request, jsonify, render_template_string, send_file
from flask_cors import CORS
from flask_socketio import SocketIO, emit
import datetime
import os
import tempfile
import warnings
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
warnings.filterwarnings('ignore')

# Audio processing imports for human detection
import numpy as np
import torch
import torchaudio
import librosa

# Add local FFmpeg to PATH if present
ffmpeg_path = r"C:\Users\niran\OneDrive\Desktop\nvenv"
if os.path.exists(os.path.join(ffmpeg_path, "ffmpeg.exe")):
    os.environ["PATH"] += os.pathsep + ffmpeg_path
    print(f"  Found local FFmpeg, added to PATH: {ffmpeg_path}")

app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

# Multi-bot data storage
bot_registry = {}  # {device_id: {"tag": "01", "latest": {...}, "history": [...]}}
bot_tag_counter = 1  # Start at 1, reserve 01 for default bot
MAX_HISTORY_PER_BOT = 500  # Max data points to keep per bot for export

# Default device patterns that get tag #01
DEFAULT_BOT_PATTERNS = ['UNKNOWN', 'ESP32', 'ESP32_BOT', 'DEFAULT']

def get_or_create_bot(device_id):
    """Get existing bot or create new one with a tag.
    - Tag #01 is reserved for default/original ESP32 bot
    - Other bots get sequential tags starting from #02
    """
    global bot_tag_counter
    
    if device_id not in bot_registry:
        # Check if this is the default/original bot (should get tag 01)
        is_default_bot = (
            device_id.upper() in DEFAULT_BOT_PATTERNS or
            device_id.upper() == 'ESP32_BOT_01' or
            not device_id or
            device_id == 'ESP32'
        )
        
        if is_default_bot:
            tag = "01"
        else:
            bot_tag_counter += 1
            tag = f"{bot_tag_counter:02d}"
        
        bot_registry[device_id] = {
            "tag": tag,
            "latest": None,
            "history": []
        }
        print(f"🤖 New bot registered: {device_id} → Tag #{tag}")
    return bot_registry[device_id]

# ============================================
# HUMAN DETECTION (Silero VAD) FUNCTIONS
# ============================================

# Global model variable (loaded once)
_vad_model = None
_vad_utils = None

def load_silero_vad():
    """Load Silero VAD model from torch.hub."""
    global _vad_model, _vad_utils
    
    if _vad_model is None:
        print("  Loading Silero VAD model...")
        model, utils = torch.hub.load(
            repo_or_dir='snakers4/silero-vad',
            model='silero_vad',
            force_reload=True,
            trust_repo=True
        )
        _vad_model = model
        _vad_utils = utils
        print("  Model loaded successfully!")
    
    return _vad_model, _vad_utils

def calculate_audio_energy(audio_data):
    """Calculate the energy/loudness of an audio segment."""
    if len(audio_data) == 0:
        return 0.0
    if len(audio_data.shape) > 1:
        audio_data = np.mean(audio_data, axis=1)
    rms = np.sqrt(np.mean(audio_data ** 2))
    return float(rms)

def auto_detect_threshold(audio_data, sr):
    """Auto-detect the audio energy threshold based on the audio content."""
    chunk_samples = int(0.1 * sr)
    if len(audio_data) < chunk_samples * 2:
        return 0.01
    
    energies = []
    for i in range(0, len(audio_data) - chunk_samples, chunk_samples):
        chunk = audio_data[i:i + chunk_samples]
        if len(chunk) > 0:
            energies.append(calculate_audio_energy(chunk))
    
    if len(energies) == 0:
        return 0.01
    
    mean_energy = np.mean(energies)
    std_energy = np.std(energies)
    threshold = max(mean_energy - (std_energy * 0.5), 0.005)
    return threshold

def detect_human_voice(audio_data, sr):
    """Detect human voice using Silero VAD deep learning model."""
    model, utils = load_silero_vad()
    (get_speech_timestamps, _, read_audio, _, _) = utils
    
    if len(audio_data.shape) > 1:
        audio_data = np.mean(audio_data, axis=1)
    
    energy = calculate_audio_energy(audio_data)
    threshold = auto_detect_threshold(audio_data, sr)
    high_energy = energy > threshold
    
    if not high_energy:
        return False, 0.0, False
    
    target_sr = 16000
    if sr != target_sr:
        audio_resampled = librosa.resample(audio_data, orig_sr=sr, target_sr=target_sr)
    else:
        audio_resampled = audio_data
    
    max_val = np.max(np.abs(audio_resampled))
    if max_val > 0:
        audio_resampled = audio_resampled / max_val
    
    audio_tensor = torch.from_numpy(audio_resampled).float()
    
    try:
        speech_timestamps = get_speech_timestamps(
            audio_tensor, model, sampling_rate=target_sr,
            threshold=0.2, min_speech_duration_ms=50, min_silence_duration_ms=50
        )
    except Exception as e:
        print(f"  Warning: VAD error - {str(e)}")
        return False, 0.0, high_energy
    
    total_samples = len(audio_resampled)
    speech_samples = sum(ts['end'] - ts['start'] for ts in speech_timestamps)
    
    if total_samples > 0:
        speech_ratio = speech_samples / total_samples
        confidence = speech_ratio * 100
    else:
        confidence = 0.0
    
    # Check for siren-like patterns
    is_siren_like = False
    if confidence < 20.0:
        spectral_centroid = librosa.feature.spectral_centroid(y=audio_resampled, sr=target_sr)[0]
        centroid_std = np.std(spectral_centroid)
        centroid_mean = np.mean(spectral_centroid)
        if centroid_mean > 400 and centroid_std > 100:
            centroid_diff = np.diff(spectral_centroid)
            smoothness = np.mean(np.abs(np.diff(centroid_diff)))
            if smoothness < 50:
                is_siren_like = True
    
    if len(speech_timestamps) > 0 and speech_ratio > 0.01 and not is_siren_like:
        return True, confidence, high_energy
    
    if high_energy and not is_siren_like and speech_ratio > 0.005:
        return True, max(confidence, 5.0), high_energy
    
    return False, confidence, high_energy

def load_audio_file(file_path):
    """Load audio file in any supported format."""
    errors = []
    
    # Try with librosa first
    try:
        audio_data, sr = librosa.load(file_path, sr=16000, mono=True)
        return audio_data, sr
    except Exception as e:
        errors.append(f"librosa: {str(e)}")
    
    # Try with torchaudio as fallback
    try:
        waveform, sr = torchaudio.load(file_path)
        if waveform.shape[0] > 1:
            waveform = torch.mean(waveform, dim=0, keepdim=True)
        if sr != 16000:
            resampler = torchaudio.transforms.Resample(sr, 16000)
            waveform = resampler(waveform)
            sr = 16000
        audio_data = waveform.squeeze().numpy()
        return audio_data, sr
    except Exception as e:
        errors.append(f"torchaudio: {str(e)}")
    
    # Log all errors
    print(f"  ❌ Failed to load audio. Errors:")
    for err in errors:
        print(f"     - {err}")
    
    return None, None

# ============================================
# API ENDPOINTS
# ============================================

# Serve the HTML dashboard
@app.route('/')
def serve_dashboard():
    html_path = os.path.join(os.path.dirname(__file__), 'lora-bot-monitor.html')
    with open(html_path, 'r', encoding='utf-8') as f:
        return f.read()

# ESP32 data endpoint
@app.route('/api/data', methods=['POST'])
def receive_data():
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"status": "error", "message": "No data received"}), 400

        device_id = data.get('device_id', 'UNKNOWN')
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data['timestamp'] = timestamp

        # Get or create bot entry
        bot = get_or_create_bot(device_id)
        data['bot_tag'] = bot['tag']  # Add tag to data
        
        # Store latest and add to history
        bot['latest'] = data
        bot['history'].append(data.copy())
        
        # Trim history if too long
        if len(bot['history']) > MAX_HISTORY_PER_BOT:
            bot['history'] = bot['history'][-MAX_HISTORY_PER_BOT:]

        print("\n" + "="*40)
        print(f"📡 DATA RECEIVED AT: {timestamp}")
        print("="*40)
        
        print(f"🆔 DEVICE ID:   {device_id}")
        print(f"🏷️  BOT TAG:     #{bot['tag']}")
        
        print(f"\n🌍 LOCATION")
        if data.get('gps_valid') == "true":
            print(f"   • Latitude:  {data.get('lat')}")
            print(f"   • Longitude: {data.get('lng')}")
        else:
            print("   • Status:    SEARCHING SATELLITES...")

        print(f"\n⚠️  ENVIRONMENT")
        print(f"   • Gas (CO):  {data.get('gas_co_ppm')} ppm")
        print(f"   • Temp:      {data.get('temperature')} °C")
        print(f"   • Pressure:  {data.get('pressure')} hPa")
        print(f"   • Altitude:  {data.get('altitude')} m")

        print(f"\n📉 MOTION & AUDIO")
        print(f"   • Accel Z:   {data.get('accel_z')} m/s²")
        print(f"   • Audio RMS: {data.get('audio_rms')} (Level)")

        print("="*40 + "\n")

        # Build bot list for frontend
        bot_list = [{"device_id": did, "tag": b['tag']} for did, b in bot_registry.items()]
        
        # Broadcast to all clients
        socketio.emit('sensor_data', {
            'data': data,
            'bot_list': bot_list,
            'active_bots': len(bot_registry)
        })
        print(f"📤 Data broadcasted (Active bots: {len(bot_registry)})")

        return jsonify({"status": "success", "message": "Data received", "bot_tag": bot['tag']}), 200

    except Exception as e:
        print(f"❌ Error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

# Human Detection Endpoint
@app.route('/api/detect-human', methods=['POST'])
def detect_human():
    """Analyze uploaded audio file for human voice."""
    try:
        if 'audio' not in request.files:
            return jsonify({"status": "error", "message": "No audio file provided"}), 400
        
        audio_file = request.files['audio']
        if audio_file.filename == '':
            return jsonify({"status": "error", "message": "No file selected"}), 400
        
        # Save to temp file
        ext = os.path.splitext(audio_file.filename)[1].lower()
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            tmp_path = tmp.name
        
        audio_file.save(tmp_path)
        
        print(f"\n{'='*40}")
        print(f"🎤 ANALYZING AUDIO: {audio_file.filename}")
        print('='*40)
        
        # Load and analyze audio
        audio_data, sr = load_audio_file(tmp_path)
        
        # Clean up temp file
        os.unlink(tmp_path)
        
        if audio_data is None:
            return jsonify({
                "status": "error", 
                "message": "Could not load audio file. Supported formats: WAV, MP3, FLAC, OGG, M4A"
            }), 400
        
        # Get duration
        duration = len(audio_data) / sr
        
        # Detect human voice
        is_human, confidence, high_energy = detect_human_voice(audio_data, sr)
        
        result = {
            "status": "success",
            "filename": audio_file.filename,
            "duration_seconds": round(duration, 2),
            "is_human_detected": bool(is_human),
            "confidence": round(float(confidence), 1),
            "high_energy": bool(high_energy)
        }
        
        # Log result
        if is_human:
            print(f"  ✅ HUMAN DETECTED (Confidence: {confidence:.1f}%)")
        else:
            print(f"  ❌ NO HUMAN DETECTED (Confidence: {confidence:.1f}%)")
        print('='*40 + "\n")
        
        return jsonify(result), 200
        
    except Exception as e:
        print(f"❌ Detection Error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

# WebSocket events
@socketio.on('connect')
def handle_connect():
    print("🔌 New WebSocket client connected!")
    # Send all bot data on connect
    if bot_registry:
        bot_list = [{"device_id": did, "tag": b['tag']} for did, b in bot_registry.items()]
        all_latest = {did: b['latest'] for did, b in bot_registry.items() if b['latest']}
        emit('init_data', {
            'bot_list': bot_list,
            'all_latest': all_latest,
            'active_bots': len(bot_registry)
        })

@socketio.on('disconnect')
def handle_disconnect():
    print("🔌 WebSocket client disconnected")

# API endpoint to get all bots and their latest data
@app.route('/api/bots', methods=['GET'])
def get_bots():
    """Get list of all registered bots with their latest data."""
    result = []
    for device_id, bot in bot_registry.items():
        result.append({
            "device_id": device_id,
            "tag": bot['tag'],
            "latest": bot['latest'],
            "history_count": len(bot['history'])
        })
    return jsonify({"bots": result, "count": len(result)})

# API endpoint to get latest data for a specific bot
@app.route('/api/latest', methods=['GET'])
@app.route('/api/latest/<device_id>', methods=['GET'])
def get_latest(device_id=None):
    if device_id:
        if device_id in bot_registry:
            return jsonify(bot_registry[device_id]['latest'])
        return jsonify({"status": "error", "message": "Bot not found"}), 404
    
    # Return all latest data if no device_id specified
    if bot_registry:
        all_latest = {did: b['latest'] for did, b in bot_registry.items() if b['latest']}
        return jsonify({"bots": all_latest, "count": len(all_latest)})
    return jsonify({"status": "waiting", "message": "No data yet"}), 200

# XLSX Export Endpoint
@app.route('/api/export', methods=['GET'])
def export_data():
    """Export all bot data to XLSX file."""
    try:
        device_id = request.args.get('device_id', None)  # Optional: export specific bot
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bot Sensor Data"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["Bot Tag", "Device ID", "Timestamp", "Temperature (°C)", "Pressure (hPa)", 
                   "Altitude (m)", "Gas CO (ppm)", "Accel Z (m/s²)", "Audio RMS", 
                   "GPS Valid", "Latitude", "Longitude"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        row_num = 2
        bots_to_export = [device_id] if device_id and device_id in bot_registry else bot_registry.keys()
        
        for did in bots_to_export:
            bot = bot_registry.get(did)
            if not bot:
                continue
            for data in bot['history']:
                ws.cell(row=row_num, column=1, value=bot['tag']).border = thin_border
                ws.cell(row=row_num, column=2, value=did).border = thin_border
                ws.cell(row=row_num, column=3, value=data.get('timestamp', '')).border = thin_border
                ws.cell(row=row_num, column=4, value=float(data.get('temperature', 0))).border = thin_border
                ws.cell(row=row_num, column=5, value=float(data.get('pressure', 0))).border = thin_border
                ws.cell(row=row_num, column=6, value=float(data.get('altitude', 0))).border = thin_border
                ws.cell(row=row_num, column=7, value=float(data.get('gas_co_ppm', 0))).border = thin_border
                ws.cell(row=row_num, column=8, value=float(data.get('accel_z', 0))).border = thin_border
                ws.cell(row=row_num, column=9, value=float(data.get('audio_rms', 0))).border = thin_border
                ws.cell(row=row_num, column=10, value=data.get('gps_valid', 'false')).border = thin_border
                ws.cell(row=row_num, column=11, value=float(data.get('lat', 0)) if data.get('lat') else '').border = thin_border
                ws.cell(row=row_num, column=12, value=float(data.get('lng', 0)) if data.get('lng') else '').border = thin_border
                row_num += 1
        
        # Adjust column widths
        column_widths = [10, 20, 22, 15, 15, 12, 14, 15, 12, 10, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"bot_data_export_{timestamp}.xlsx"
        
        print(f"📊 Exported {row_num - 2} data points to {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"❌ Export Error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == '__main__':
    print(f"🚀 Server running on port 5000...")
    print(f"   Dashboard: http://localhost:5000")
    print(f"   Multi-Bot API: GET /api/bots")
    print(f"   Export API: GET /api/export")
    print(f"   Human Detection: POST /api/detect-human")
    print(f"   WebSocket ready for real-time updates!")
    socketio.run(app, host='0.0.0.0', port=5000, debug=True, allow_unsafe_werkzeug=True)